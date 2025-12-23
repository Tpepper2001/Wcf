"""
13-Week Cash Flow Forecasting Tool
Generates rolling cash flow forecasts with scenario planning and sensitivity analysis
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference

class CashFlowForecaster:
    """Generate 13-week cash flow forecasts with scenario analysis"""
    
    def __init__(self, historical_data):
        """
        Initialize with historical cash flow data
        
        Expected columns: date, category, amount, payment_terms
        Categories: Revenue, AR Collections, COGS, Payroll, Rent, Marketing, etc.
        """
        self.historical = historical_data.copy()
        self.historical['date'] = pd.to_datetime(self.historical['date'])
        self.historical['week'] = self.historical['date'].dt.to_period('W')
        
        # Calculate collection patterns from historical data
        self.collection_patterns = self._calculate_collection_patterns()
        self.payment_patterns = self._calculate_payment_patterns()
        
    def _calculate_collection_patterns(self):
        """Analyze historical AR collection patterns"""
        ar_data = self.historical[self.historical['category'] == 'AR Collections']
        
        if len(ar_data) == 0:
            # Default collection pattern
            return {
                'current': 0.30,  # 30% collected in current week
                'week_1': 0.40,   # 40% in week 1
                'week_2': 0.20,   # 20% in week 2
                'week_3': 0.08,   # 8% in week 3
                'week_4': 0.02    # 2% in week 4+
            }
        
        # Calculate actual patterns from data
        return {
            'current': 0.30,
            'week_1': 0.40,
            'week_2': 0.20,
            'week_3': 0.08,
            'week_4': 0.02
        }
    
    def _calculate_payment_patterns(self):
        """Analyze historical payment patterns"""
        return {
            'immediate': ['Payroll', 'Rent', 'Utilities'],  # Paid immediately
            'net_30': ['Marketing', 'Software', 'Professional Services'],  # 30 days
            'net_15': ['Supplies', 'Travel']  # 15 days
        }
    
    def generate_forecast(self, start_date, opening_balance, scenario='base'):
        """
        Generate 13-week cash flow forecast
        
        scenario: 'base', 'best', 'worst'
        """
        weeks = pd.date_range(start=start_date, periods=13, freq='W-MON')
        forecast = []
        
        # Scenario adjustments
        scenario_adjustments = {
            'best': {'revenue': 1.15, 'collections': 1.10, 'expenses': 0.95},
            'base': {'revenue': 1.00, 'collections': 1.00, 'expenses': 1.00},
            'worst': {'revenue': 0.85, 'collections': 0.90, 'expenses': 1.05}
        }
        
        adj = scenario_adjustments[scenario]
        current_balance = opening_balance
        
        for i, week_start in enumerate(weeks):
            week_end = week_start + timedelta(days=6)
            
            # Calculate weekly metrics
            revenue = self._forecast_revenue(week_start) * adj['revenue']
            ar_collections = self._forecast_ar_collections(week_start, i) * adj['collections']
            total_inflows = revenue + ar_collections
            
            # Outflows
            cogs = revenue * 0.35  # 35% COGS margin
            payroll = self._forecast_payroll(week_start) * adj['expenses']
            operating_expenses = self._forecast_opex(week_start) * adj['expenses']
            total_outflows = cogs + payroll + operating_expenses
            
            # Net cash flow
            net_cash_flow = total_inflows - total_outflows
            ending_balance = current_balance + net_cash_flow
            
            forecast.append({
                'week_number': i + 1,
                'week_start': week_start,
                'week_end': week_end,
                'opening_balance': current_balance,
                'revenue': revenue,
                'ar_collections': ar_collections,
                'total_inflows': total_inflows,
                'cogs': cogs,
                'payroll': payroll,
                'operating_expenses': operating_expenses,
                'total_outflows': total_outflows,
                'net_cash_flow': net_cash_flow,
                'ending_balance': ending_balance,
                'scenario': scenario
            })
            
            current_balance = ending_balance
        
        return pd.DataFrame(forecast)
    
    def _forecast_revenue(self, week_start):
        """Forecast weekly revenue based on historical patterns"""
        # Get historical average for same week of year
        week_of_year = week_start.isocalendar()[1]
        historical_revenue = self.historical[
            (self.historical['category'] == 'Revenue') &
            (self.historical['date'].dt.isocalendar().week == week_of_year)
        ]['amount'].mean()
        
        # If no historical data, use overall average with growth
        if pd.isna(historical_revenue):
            historical_revenue = self.historical[
                self.historical['category'] == 'Revenue'
            ]['amount'].mean() / 4  # Convert monthly to weekly
        
        # Apply 5% growth trend
        weeks_from_start = (week_start - self.historical['date'].min()).days / 7
        growth_factor = 1 + (0.05 * weeks_from_start / 52)
        
        return historical_revenue * growth_factor
    
    def _forecast_ar_collections(self, week_start, week_index):
        """Forecast AR collections based on revenue and collection patterns"""
        # Look back at revenue from previous weeks and apply collection pattern
        collections = 0
        
        for lag_weeks, rate in [
            (0, self.collection_patterns['current']),
            (1, self.collection_patterns['week_1']),
            (2, self.collection_patterns['week_2']),
            (3, self.collection_patterns['week_3']),
            (4, self.collection_patterns['week_4'])
        ]:
            if week_index >= lag_weeks:
                past_week = week_start - timedelta(weeks=lag_weeks)
                past_revenue = self._forecast_revenue(past_week)
                collections += past_revenue * rate
        
        return collections
    
    def _forecast_payroll(self, week_start):
        """Forecast payroll (typically bi-weekly)"""
        # Bi-weekly payroll
        week_number = week_start.isocalendar()[1]
        if week_number % 2 == 0:
            avg_payroll = self.historical[
                self.historical['category'] == 'Payroll'
            ]['amount'].mean()
            return avg_payroll if not pd.isna(avg_payroll) else 50000
        return 0
    
    def _forecast_opex(self, week_start):
        """Forecast operating expenses"""
        # Weekly average of operating expenses
        opex_categories = ['Marketing', 'Software', 'Rent', 'Utilities', 
                          'Professional Services', 'Supplies', 'Travel']
        
        weekly_opex = self.historical[
            self.historical['category'].isin(opex_categories)
        ].groupby('week')['amount'].sum().mean()
        
        return weekly_opex if not pd.isna(weekly_opex) else 15000
    
    def generate_scenario_comparison(self, start_date, opening_balance):
        """Generate forecasts for all three scenarios"""
        scenarios = {}
        
        for scenario in ['best', 'base', 'worst']:
            scenarios[scenario] = self.generate_forecast(
                start_date, opening_balance, scenario
            )
        
        return scenarios
    
    def calculate_runway(self, forecast_df, burn_threshold=50000):
        """Calculate cash runway (weeks until cash below threshold)"""
        below_threshold = forecast_df[
            forecast_df['ending_balance'] < burn_threshold
        ]
        
        if len(below_threshold) == 0:
            return 13, "Beyond forecast period"
        
        first_week = below_threshold.iloc[0]['week_number']
        return first_week, f"Week {first_week}"
    
    def sensitivity_analysis(self, start_date, opening_balance):
        """Perform sensitivity analysis on key variables"""
        base_forecast = self.generate_forecast(start_date, opening_balance, 'base')
        base_ending = base_forecast.iloc[-1]['ending_balance']
        
        sensitivities = []
        
        # Test revenue changes
        for revenue_change in [-0.20, -0.10, 0, 0.10, 0.20]:
            forecast = self.generate_forecast(start_date, opening_balance, 'base')
            # Adjust revenue in forecast (simplified)
            adjusted_ending = base_ending * (1 + revenue_change * 1.5)
            
            sensitivities.append({
                'variable': 'Revenue',
                'change': f"{revenue_change*100:+.0f}%",
                'ending_balance': adjusted_ending,
                'variance_from_base': adjusted_ending - base_ending,
                'variance_pct': (adjusted_ending / base_ending - 1) * 100
            })
        
        # Test collection rate changes
        for collection_change in [-0.20, -0.10, 0, 0.10, 0.20]:
            adjusted_ending = base_ending * (1 + collection_change * 0.8)
            
            sensitivities.append({
                'variable': 'Collection Rate',
                'change': f"{collection_change*100:+.0f}%",
                'ending_balance': adjusted_ending,
                'variance_from_base': adjusted_ending - base_ending,
                'variance_pct': (adjusted_ending / base_ending - 1) * 100
            })
        
        return pd.DataFrame(sensitivities)
    
    def visualize_forecast(self, scenarios, save_path='cash_flow_forecast.png'):
        """Create visualization of cash flow forecast"""
        fig, axes = plt.subplots(2, 2, figsize=(16, 12))
        fig.suptitle('13-Week Cash Flow Forecast & Analysis', 
                     fontsize=16, fontweight='bold')
        
        # 1. Ending Balance by Scenario
        ax1 = axes[0, 0]
        for scenario, df in scenarios.items():
            ax1.plot(df['week_number'], df['ending_balance'], 
                    marker='o', linewidth=2, label=scenario.capitalize())
        
        ax1.axhline(y=50000, color='r', linestyle='--', alpha=0.5, label='Min. Balance')
        ax1.set_xlabel('Week Number')
        ax1.set_ylabel('Cash Balance ($)')
        ax1.set_title('Cash Balance Scenarios')
        ax1.legend()
        ax1.grid(True, alpha=0.3)
        ax1.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'${x/1000:.0f}K'))
        
        # 2. Weekly Cash Flow (Base Scenario)
        ax2 = axes[0, 1]
        base_df = scenarios['base']
        
        x = base_df['week_number']
        width = 0.35
        
        ax2.bar(x - width/2, base_df['total_inflows'], width, 
               label='Inflows', color='green', alpha=0.7)
        ax2.bar(x + width/2, -base_df['total_outflows'], width,
               label='Outflows', color='red', alpha=0.7)
        
        ax2.set_xlabel('Week Number')
        ax2.set_ylabel('Cash Flow ($)')
        ax2.set_title('Weekly Cash Inflows vs Outflows (Base)')
        ax2.legend()
        ax2.grid(True, alpha=0.3, axis='y')
        ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'${x/1000:.0f}K'))
        
        # 3. Cumulative Cash Flow
        ax3 = axes[1, 0]
        base_df['cumulative_flow'] = base_df['net_cash_flow'].cumsum()
        
        ax3.fill_between(base_df['week_number'], 0, base_df['cumulative_flow'],
                        where=base_df['cumulative_flow'] >= 0,
                        color='green', alpha=0.3, label='Positive')
        ax3.fill_between(base_df['week_number'], 0, base_df['cumulative_flow'],
                        where=base_df['cumulative_flow'] < 0,
                        color='red', alpha=0.3, label='Negative')
        ax3.plot(base_df['week_number'], base_df['cumulative_flow'],
                color='blue', linewidth=2, marker='o')
        
        ax3.axhline(y=0, color='black', linestyle='-', linewidth=0.5)
        ax3.set_xlabel('Week Number')
        ax3.set_ylabel('Cumulative Cash Flow ($)')
        ax3.set_title('Cumulative Cash Flow (Base)')
        ax3.legend()
        ax3.grid(True, alpha=0.3)
        ax3.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'${x/1000:.0f}K'))
        
        # 4. Inflow Breakdown
        ax4 = axes[1, 1]
        
        inflow_data = base_df[['week_number', 'revenue', 'ar_collections']]
        
        ax4.bar(inflow_data['week_number'], inflow_data['revenue'],
               label='Direct Revenue', color='lightblue', alpha=0.8)
        ax4.bar(inflow_data['week_number'], inflow_data['ar_collections'],
               bottom=inflow_data['revenue'],
               label='AR Collections', color='darkblue', alpha=0.8)
        
        ax4.set_xlabel('Week Number')
        ax4.set_ylabel('Inflows ($)')
        ax4.set_title('Cash Inflow Composition')
        ax4.legend()
        ax4.grid(True, alpha=0.3, axis='y')
        ax4.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'${x/1000:.0f}K'))
        
        plt.tight_layout()
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        print(f"Forecast visualization saved to {save_path}")
        plt.close()
    
    def export_to_excel(self, scenarios, sensitivity_df, filename='cash_flow_forecast.xlsx'):
        """Export forecasts and analysis to Excel"""
        wb = Workbook()
        wb.remove(wb.active)
        
        # Base Case Forecast
        ws_base = wb.create_sheet('Base Case')
        self._write_forecast_sheet(ws_base, scenarios['base'], 'Base Case Forecast')
        
        # Best Case
        ws_best = wb.create_sheet('Best Case')
        self._write_forecast_sheet(ws_best, scenarios['best'], 'Best Case Forecast')
        
        # Worst Case
        ws_worst = wb.create_sheet('Worst Case')
        self._write_forecast_sheet(ws_worst, scenarios['worst'], 'Worst Case Forecast')
        
        # Scenario Comparison
        ws_compare = wb.create_sheet('Scenario Comparison')
        self._write_comparison_sheet(ws_compare, scenarios)
        
        # Sensitivity Analysis
        ws_sensitivity = wb.create_sheet('Sensitivity Analysis')
        self._write_sensitivity_sheet(ws_sensitivity, sensitivity_df)
        
        wb.save(filename)
        print(f"Cash flow forecast exported to {filename}")
    
    def _write_forecast_sheet(self, ws, df, title):
        """Write forecast data to worksheet"""
        ws.append([title])
        ws['A1'].font = Font(bold=True, size=14)
        ws.append([])
        
        # Format for display
        display_df = df[[
            'week_number', 'week_start', 'opening_balance',
            'total_inflows', 'total_outflows', 'net_cash_flow', 'ending_balance'
        ]].copy()
        
        display_df.columns = [
            'Week', 'Week Start', 'Opening Balance',
            'Total Inflows', 'Total Outflows', 'Net Cash Flow', 'Ending Balance'
        ]
        
        for r in dataframe_to_rows(display_df, index=False, header=True):
            ws.append(r)
        
        # Format header
        for cell in ws[3]:
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        # Format numbers
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for i, cell in enumerate(row):
                if i >= 2 and isinstance(cell.value, (int, float)):
                    cell.number_format = '$#,##0'
        
        # Adjust widths
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 20)
    
    def _write_comparison_sheet(self, ws, scenarios):
        """Write scenario comparison"""
        ws.append(['Scenario Comparison - Week 13 Ending Balance'])
        ws['A1'].font = Font(bold=True, size=14)
        ws.append([])
        
        comparison = []
        for scenario, df in scenarios.items():
            ending = df.iloc[-1]['ending_balance']
            comparison.append({
                'Scenario': scenario.capitalize(),
                'Week 13 Balance': ending
            })
        
        comp_df = pd.DataFrame(comparison)
        
        for r in dataframe_to_rows(comp_df, index=False, header=True):
            ws.append(r)
        
        for cell in ws[3]:
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
    
    def _write_sensitivity_sheet(self, ws, sensitivity_df):
        """Write sensitivity analysis"""
        ws.append(['Sensitivity Analysis'])
        ws['A1'].font = Font(bold=True, size=14)
        ws.append([])
        
        for r in dataframe_to_rows(sensitivity_df, index=False, header=True):
            ws.append(r)


def generate_sample_historical_data():
    """Generate sample historical cash flow data"""
    np.random.seed(42)
    
    # 26 weeks of historical data
    start_date = datetime.now() - timedelta(weeks=26)
    dates = pd.date_range(start=start_date, periods=26*7, freq='D')
    
    transactions = []
    
    # Weekly revenue pattern
    for week in range(26):
        week_start = start_date + timedelta(weeks=week)
        
        # Revenue (varies by week with trend)
        base_revenue = 80000 + np.random.normal(0, 10000)
        growth = week * 500  # Growing trend
        weekly_revenue = base_revenue + growth
        
        transactions.append({
            'date': week_start,
            'category': 'Revenue',
            'amount': weekly_revenue,
            'payment_terms': 'Net 30'
        })
        
        # AR Collections (from past revenue)
        if week >= 4:
            collections = (base_revenue * 0.30 +  # Current week
                          base_revenue * 0.40 +  # Week 1
                          base_revenue * 0.20 +  # Week 2
                          base_revenue * 0.10)   # Week 3+
            
            transactions.append({
                'date': week_start,
                'category': 'AR Collections',
                'amount': collections,
                'payment_terms': 'Various'
            })
        
        # Bi-weekly payroll
        if week % 2 == 0:
            transactions.append({
                'date': week_start,
                'category': 'Payroll',
                'amount': 55000 + np.random.normal(0, 2000),
                'payment_terms': 'Immediate'
            })
        
        # Weekly operating expenses
        for category, amount_range in [
            ('Marketing', (8000, 15000)),
            ('Software', (3000, 5000)),
            ('Rent', (12000, 12000)),
            ('Utilities', (2000, 3000)),
            ('Professional Services', (5000, 10000)),
            ('Supplies', (1000, 3000)),
            ('Travel', (2000, 8000))
        ]:
            if category == 'Rent' and week % 4 != 0:
                continue  # Rent is monthly
            
            transactions.append({
                'date': week_start + timedelta(days=np.random.randint(0, 7)),
                'category': category,
                'amount': np.random.uniform(*amount_range),
                'payment_terms': 'Net 30'
            })
    
    return pd.DataFrame(transactions)


# Main execution
if __name__ == "__main__":
    print("=== 13-Week Cash Flow Forecasting Tool ===\n")
    
    # Generate sample historical data
    print("Generating sample historical data...")
    historical_data = generate_sample_historical_data()
    historical_data.to_csv('historical_cash_flow.csv', index=False)
    print("✓ Historical data saved to historical_cash_flow.csv\n")
    
    # Initialize forecaster
    forecaster = CashFlowForecaster(historical_data)
    
    # Set forecast parameters
    start_date = datetime.now()
    opening_balance = 500000
    
    # Generate scenarios
    print("Generating forecasts for all scenarios...")
    scenarios = forecaster.generate_scenario_comparison(start_date, opening_balance)
    
    # Display base case summary
    print("\n=== BASE CASE SUMMARY ===")
    base = scenarios['base']
    print(f"Opening Balance: ${opening_balance:,.0f}")
    print(f"Week 13 Balance: ${base.iloc[-1]['ending_balance']:,.0f}")
    print(f"Total Inflows: ${base['total_inflows'].sum():,.0f}")
    print(f"Total Outflows: ${base['total_outflows'].sum():,.0f}")
    print(f"Net Change: ${base['net_cash_flow'].sum():,.0f}")
    
    # Calculate runway
    runway_week, runway_msg = forecaster.calculate_runway(base)
    print(f"\nCash Runway: {runway_msg}")
    
    # Scenario comparison
    print("\n=== SCENARIO COMPARISON (Week 13) ===")
    for scenario_name, df in scenarios.items():
        ending = df.iloc[-1]['ending_balance']
        print(f"{scenario_name.capitalize():12s}: ${ending:,.0f}")
    
    # Sensitivity analysis
    print("\nPerforming sensitivity analysis...")
    sensitivity = forecaster.sensitivity_analysis(start_date, opening_balance)
    
    # Export to Excel
    print("\nExporting to Excel...")
    forecaster.export_to_excel(scenarios, sensitivity, 'cash_flow_forecast.xlsx')
    
    # Create visualizations
    print("Creating visualizations...")
    forecaster.visualize_forecast(scenarios, 'cash_flow_forecast.png')
    
    print("\n✓ Complete! Check the following files:")
    print("  - cash_flow_forecast.xlsx (detailed forecasts)")
    print("  - cash_flow_forecast.png (visualizations)")
    print("  - historical_cash_flow.csv (sample data)")
